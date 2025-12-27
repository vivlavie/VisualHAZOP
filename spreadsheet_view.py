"""
Spreadsheet view for displaying nodes and deviations.
"""
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from models import Node, Deviation, HAZOPData
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


class SpreadsheetView(tk.Toplevel):
    """Window for displaying nodes and deviations in a spreadsheet format."""
    
    def __init__(self, parent, hazop_data: HAZOPData):
        super().__init__(parent)
        self.parent_app = parent  # Store reference to main app
        self.hazop_data = hazop_data
        self.title("HAZOP Analysis Spreadsheet")
        self.geometry("1200x600")
        
        # Mapping from treeview item IDs to (node, deviation) tuples
        self.item_to_deviation = {}
        
        self.create_widgets()
        self.refresh_data()
    
    def create_widgets(self):
        """Create the spreadsheet widgets."""
        # Toolbar
        toolbar = ttk.Frame(self)
        toolbar.pack(fill=tk.X, padx=5, pady=5)
        
        ttk.Button(toolbar, text="Refresh", command=self.refresh_data).pack(side=tk.LEFT, padx=2)
        ttk.Button(toolbar, text="Export to Excel", command=self.export_to_excel).pack(side=tk.LEFT, padx=2)
        
        # Treeview with scrollbars
        frame = ttk.Frame(self)
        frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        # Create treeview
        columns = ("Deviation", "Causes", "Consequence", "Safeguards", "Recommendations")
        self.tree = ttk.Treeview(frame, columns=columns, show="tree headings", height=20)
        
        # Configure columns
        self.tree.heading("#0", text="Node")
        self.tree.column("#0", width=150)
        
        for col in columns:
            self.tree.heading(col, text=col)
            self.tree.column(col, width=200)
        
        # Scrollbars
        v_scrollbar = ttk.Scrollbar(frame, orient=tk.VERTICAL, command=self.tree.yview)
        h_scrollbar = ttk.Scrollbar(frame, orient=tk.HORIZONTAL, command=self.tree.xview)
        self.tree.configure(yscrollcommand=v_scrollbar.set, xscrollcommand=h_scrollbar.set)
        
        # Bind double-click event
        self.tree.bind("<Double-Button-1>", self.on_double_click)
        
        # Pack
        self.tree.grid(row=0, column=0, sticky=tk.NSEW)
        v_scrollbar.grid(row=0, column=1, sticky=tk.NS)
        h_scrollbar.grid(row=1, column=0, sticky=tk.EW)
        
        frame.grid_rowconfigure(0, weight=1)
        frame.grid_columnconfigure(0, weight=1)
    
    def refresh_data(self):
        """Refresh the spreadsheet data."""
        # Clear existing items
        for item in self.tree.get_children():
            self.tree.delete(item)
        
        # Check if we have data
        if not self.hazop_data or not self.hazop_data.nodes:
            # Show empty message
            self.tree.insert("", tk.END, text="No nodes found", values=("", "", "", "", ""))
            return
        
        # Group nodes by page
        nodes_by_page = {}
        for node in self.hazop_data.nodes:
            page = node.page_number
            if page not in nodes_by_page:
                nodes_by_page[page] = []
            nodes_by_page[page].append(node)
        
        # Add nodes
        for page_num in sorted(nodes_by_page.keys()):
            page_nodes = nodes_by_page[page_num]
            for node in page_nodes:
                # Create node item with background color
                node_id = self.tree.insert("", tk.END, text=node.name or f"Node {page_num}",
                                          values=("", "", "", "", ""),
                                          tags=("node",))
                
                # Configure node row color
                rgb = self.hex_to_rgb(node.color)
                # Convert to a lighter shade for background
                bg_color = f"#{min(255, rgb[0]+50):02x}{min(255, rgb[1]+50):02x}{min(255, rgb[2]+50):02x}"
                text_color = "#000000" if sum(rgb) > 384 else "#FFFFFF"
                
                self.tree.set(node_id, "Deviation", f"Page {page_num + 1}")
                
                # Add deviations
                if node.deviations:
                    for dev in node.deviations:
                        # Calculate max rows needed for this deviation
                        max_items = max(
                            len(dev.causes) if dev.causes else 0,
                            len(dev.safeguards) if dev.safeguards else 0,
                            len(dev.recommendations) if dev.recommendations else 0,
                            1  # At least one row for deviation and consequence
                        )
                        
                        # Create rows for this deviation
                        # Show deviation and consequence on all rows for clarity
                        for i in range(max_items):
                            # Deviation name (show on all rows)
                            deviation_text = dev.deviation if i == 0 else "↳"  # Continuation marker
                            
                            # Causes (one per row)
                            cause_text = dev.causes[i] if dev.causes and i < len(dev.causes) else ""
                            
                            # Consequence (show on all rows)
                            consequence_text = dev.consequence if i == 0 else "↳"
                            
                            # Safeguards (one per row)
                            safeguard_text = dev.safeguards[i] if dev.safeguards and i < len(dev.safeguards) else ""
                            
                            # Recommendations (one per row)
                            recommendation_text = dev.recommendations[i] if dev.recommendations and i < len(dev.recommendations) else ""
                            
                            dev_item_id = self.tree.insert(node_id, tk.END,
                                            text="",
                                            values=(
                                                deviation_text,
                                                cause_text,
                                                consequence_text,
                                                safeguard_text,
                                                recommendation_text
                                            ),
                                            tags=("deviation",))
                            
                            # Store mapping: store for all rows of this deviation
                            # This allows double-clicking on any row to edit the deviation
                            self.item_to_deviation[dev_item_id] = (node, dev)
                else:
                    # Insert empty deviation row
                    self.tree.insert(node_id, tk.END,
                                    text="",
                                    values=("", "", "", "", ""),
                                    tags=("deviation",))
        
        # Configure tags
        self.tree.tag_configure("node", background="#E0E0E0", font=("Arial", 10, "bold"))
        self.tree.tag_configure("deviation", background="#FFFFFF")
        
        # Set row height to accommodate content
        style = ttk.Style()
        style.configure("Treeview", rowheight=30)  # Increased row height for better visibility
    
    def on_double_click(self, event):
        """Handle double-click on a treeview item."""
        item = self.tree.selection()[0] if self.tree.selection() else None
        if not item:
            return
        
        # Check if this is a deviation row
        if item in self.item_to_deviation:
            node, deviation = self.item_to_deviation[item]
            # Open deviation editor
            from deviation_editor import DeviationEditor
            
            def on_save(dev):
                # The deviation object is updated in place, just refresh the view
                self.refresh_data()
                # Notify parent app to update PDF viewer if needed
                if self.parent_app and hasattr(self.parent_app, 'pdf_viewer'):
                    self.parent_app.pdf_viewer.render_page()
            
            editor = DeviationEditor(self, deviation, on_save_callback=on_save)
            self.wait_window(editor)
    
    def hex_to_rgb(self, hex_color):
        """Convert hex color to RGB tuple."""
        hex_color = hex_color.lstrip('#')
        return tuple(int(hex_color[i:i+2], 16) for i in (0, 2, 4))
    
    def export_to_excel(self):
        """Export data to Excel file."""
        filename = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        
        if not filename:
            return
        
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "HAZOP Analysis"
            
            # Headers
            headers = ["Node", "Page", "Deviation", "Causes", "Consequence", "Safeguards", "Recommendations", "Comments"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Border style
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            row = 2
            
            # Group nodes by page
            nodes_by_page = {}
            for node in self.hazop_data.nodes:
                page = node.page_number
                if page not in nodes_by_page:
                    nodes_by_page[page] = []
                nodes_by_page[page].append(node)
            
            # Write data
            for page_num in sorted(nodes_by_page.keys()):
                page_nodes = nodes_by_page[page_num]
                for node in page_nodes:
                    node_name = node.name or f"Node {page_num + 1}"
                    
                    if node.deviations:
                        for dev in node.deviations:
                            # Calculate max rows needed for this deviation
                            max_items = max(
                                len(dev.causes),
                                len(dev.safeguards),
                                len(dev.recommendations),
                                1
                            )
                            
                            for i in range(max_items):
                                # Node (span across rows)
                                if i == 0:
                                    ws.cell(row=row, column=1, value=node_name)
                                    if max_items > 1:
                                        ws.merge_cells(f'A{row}:A{row + max_items - 1}')
                                        ws.cell(row=row, column=1).alignment = Alignment(vertical="top", wrap_text=True)
                                
                                # Page (span across rows)
                                if i == 0:
                                    ws.cell(row=row, column=2, value=page_num + 1)
                                    if max_items > 1:
                                        ws.merge_cells(f'B{row}:B{row + max_items - 1}')
                                        ws.cell(row=row, column=2).alignment = Alignment(vertical="top", wrap_text=True)
                                
                                # Deviation (span across rows)
                                if i == 0:
                                    ws.cell(row=row, column=3, value=dev.deviation)
                                    if max_items > 1:
                                        ws.merge_cells(f'C{row}:C{row + max_items - 1}')
                                        ws.cell(row=row, column=3).alignment = Alignment(vertical="top", wrap_text=True)
                                
                                # Causes
                                if i < len(dev.causes):
                                    ws.cell(row=row, column=4, value=dev.causes[i])
                                
                                # Consequence (span across rows)
                                if i == 0:
                                    ws.cell(row=row, column=5, value=dev.consequence)
                                    if max_items > 1:
                                        ws.merge_cells(f'E{row}:E{row + max_items - 1}')
                                        ws.cell(row=row, column=5).alignment = Alignment(vertical="top", wrap_text=True)
                                
                                # Safeguards
                                if i < len(dev.safeguards):
                                    ws.cell(row=row, column=6, value=dev.safeguards[i])
                                
                                # Recommendations
                                if i < len(dev.recommendations):
                                    ws.cell(row=row, column=7, value=dev.recommendations[i])
                                
                                # Comments
                                if i == 0:
                                    ws.cell(row=row, column=8, value=dev.comments)
                                    if max_items > 1:
                                        ws.merge_cells(f'H{row}:H{row + max_items - 1}')
                                        ws.cell(row=row, column=8).alignment = Alignment(vertical="top", wrap_text=True)
                                
                                # Apply borders
                                for col in range(1, 9):
                                    ws.cell(row=row, column=col).border = thin_border
                                
                                # Apply node background color to first column (merged cell)
                                if i == 0:
                                    rgb = self.hex_to_rgb(node.color)
                                    bg_color = f"{min(255, rgb[0]+50):02x}{min(255, rgb[1]+50):02x}{min(255, rgb[2]+50):02x}"
                                    fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
                                    # Apply fill to the merged cell range
                                    if max_items > 1:
                                        for r in range(row, row + max_items):
                                            ws.cell(row=r, column=1).fill = fill
                                    else:
                                        ws.cell(row=row, column=1).fill = fill
                                
                                row += 1
                    else:
                        # Node with no deviations
                        ws.cell(row=row, column=1, value=node_name)
                        ws.cell(row=row, column=2, value=page_num + 1)
                        for col in range(1, 9):
                            ws.cell(row=row, column=col).border = thin_border
                        rgb = self.hex_to_rgb(node.color)
                        bg_color = f"{min(255, rgb[0]+50):02x}{min(255, rgb[1]+50):02x}{min(255, rgb[2]+50):02x}"
                        fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
                        ws.cell(row=row, column=1).fill = fill
                        row += 1
            
            # Auto-adjust column widths
            for col in range(1, 9):
                max_length = 0
                column = get_column_letter(col)
                for cell in ws[column]:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column].width = adjusted_width
            
            wb.save(filename)
            messagebox.showinfo("Success", f"Data exported to {filename}")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to export: {str(e)}")

